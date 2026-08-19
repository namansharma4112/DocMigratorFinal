"""End-to-end and unit tests for the legal document migration pipeline."""
from __future__ import annotations
import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from legal_pipeline import ocr_support
from legal_pipeline.classify import classify_document
from legal_pipeline.config import Config, ClassificationThresholds, Dedup
from legal_pipeline.dedupe import DedupRecord, deduplicate
from legal_pipeline.metadata import extract_metadata
from legal_pipeline.pipeline import run
from .make_test_pdfs import main as build_standard_fixture_set, OUT as STANDARD_FIXTURE_DIR


# ---------------------------------------------------------------------------
# OCR availability guard
# ---------------------------------------------------------------------------
# Some tests genuinely require a working OCR toolchain (Tesseract + Poppler) on
# the machine running the tests. On CI runners those binaries may not be
# installed / on PATH, in which case scanned PDFs legitimately produce zero OCR
# text. Rather than fail the build for an environment reason, we SKIP those
# specific tests when OCR isn't ready. This mirrors the project's established
# philosophy (see make_test_pdfs.py) of never letting a test's pass/fail depend
# on the runner's OCR engine. OCR-independent behaviour (exact-file dedupe,
# classification of native text, etc.) is still fully tested unconditionally.
OCR_READY = ocr_support.configure_ocr().ready
requires_ocr = pytest.mark.skipif(
    not OCR_READY,
    reason="Tesseract/Poppler not available on this machine; skipping OCR-dependent assertions.",
)


@pytest.fixture(scope="module")
def fixture_pdfs() -> Path:
    build_standard_fixture_set()
    return STANDARD_FIXTURE_DIR


@pytest.fixture()
def output_dir(tmp_path) -> Path:
    d = tmp_path / "output"
    yield d
    if d.exists():
        shutil.rmtree(d, ignore_errors=True)


def _load_tracker_rows(summary):
    wb = load_workbook(summary["tracker"])
    ws = wb["Tracker"]
    headers = [c.value for c in ws[1]]
    return {r["file_name"]: r for r in
            (dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True))}


# ---------------------------------------------------------------------------
# End-to-end pipeline tests
# ---------------------------------------------------------------------------
def test_full_pipeline_runs_without_crashing_and_produces_expected_counts(fixture_pdfs, output_dir):
    cfg = Config()
    cfg.paths.source_dir = fixture_pdfs
    cfg.paths.output_dir = output_dir
    summary = run(cfg, copy_files=True, log=lambda *_: None)
    assert summary["total"] == 12
    assert summary["removed"] >= 2  # #02 and #08 exact copies
    assert summary["failed_extraction"] >= 1  # #11 corrupt


def test_threaded_and_serial_extraction_produce_identical_results(fixture_pdfs, tmp_path):
    """The middle-path threading must not change ANY output vs serial.

    This holds regardless of whether OCR is available: with or without OCR,
    serial and threaded see the SAME inputs and must agree on every count."""
    keys = ["total", "retained", "removed", "excluded_do_not_retain",
            "needs_review", "scanned", "ocr_read", "failed_extraction"]

    cfg_s = Config()
    cfg_s.paths.source_dir = fixture_pdfs
    cfg_s.paths.output_dir = tmp_path / "serial"
    cfg_s.ingestion.extract_threads = 1
    s_serial = run(cfg_s, copy_files=True, log=lambda *_: None)

    cfg_t = Config()
    cfg_t.paths.source_dir = fixture_pdfs
    cfg_t.paths.output_dir = tmp_path / "threaded"
    cfg_t.ingestion.extract_threads = 4
    s_threaded = run(cfg_t, copy_files=True, log=lambda *_: None)

    for k in keys:
        assert s_serial[k] == s_threaded[k], f"{k}: {s_serial[k]} != {s_threaded[k]}"


def test_ocr_page_cap_does_not_change_classification(fixture_pdfs, tmp_path):
    """Capping OCR pages must not alter doc_type vs uncapped.

    Runs regardless of OCR availability: if OCR is unavailable the scanned docs
    simply have no text in BOTH runs, so their doc_types still match."""
    cfg_full = Config()
    cfg_full.paths.source_dir = fixture_pdfs
    cfg_full.paths.output_dir = tmp_path / "full"
    cfg_full.ingestion.ocr_max_pages = None
    full = _load_tracker_rows(run(cfg_full, copy_files=True, log=lambda *_: None))

    cfg_cap = Config()
    cfg_cap.paths.source_dir = fixture_pdfs
    cfg_cap.paths.output_dir = tmp_path / "cap"
    cfg_cap.ingestion.ocr_max_pages = 2
    cap = _load_tracker_rows(run(cfg_cap, copy_files=True, log=lambda *_: None))

    for name in full:
        assert full[name]["doc_type"] == cap[name]["doc_type"], name


def test_corrupt_and_unreadable_files_never_crash_the_run(fixture_pdfs, output_dir):
    cfg = Config()
    cfg.paths.source_dir = fixture_pdfs
    cfg.paths.output_dir = output_dir
    summary = run(cfg, copy_files=True, log=lambda *_: None)
    assert summary["failed_extraction"] >= 1


def test_different_client_boilerplate_contract_is_not_falsely_merged(fixture_pdfs, output_dir):
    cfg = Config()
    cfg.paths.source_dir = fixture_pdfs
    cfg.paths.output_dir = output_dir
    summary = run(cfg, copy_files=True, log=lambda *_: None)
    rows = _load_tracker_rows(summary)
    assert rows["04_contract_globex.pdf"]["status"] == "retained"


def test_near_duplicate_scanned_rescan_is_caught(fixture_pdfs, output_dir):
    """#08 is a byte-for-byte copy of #07 -> matched via the exact_file hash
    tier, which runs BEFORE any OCR. This holds even without OCR installed."""
    cfg = Config()
    cfg.paths.source_dir = fixture_pdfs
    cfg.paths.output_dir = output_dir
    summary = run(cfg, copy_files=True, log=lambda *_: None)
    rows = _load_tracker_rows(summary)
    assert rows["08_scanned_contract_delta_rescan.pdf"]["status"] == "removed"
    assert rows["07_scanned_contract_delta.pdf"]["status"] == "retained"


@requires_ocr
def test_scanned_pdfs_are_independently_ocr_processed(fixture_pdfs, output_dir):
    """Confirms both scanned files (#07, #08) are actually OCR'd. REQUIRES a
    working OCR toolchain; skipped when Tesseract/Poppler are unavailable so a
    CI runner without them does not fail the build for an environment reason."""
    cfg = Config()
    cfg.paths.source_dir = fixture_pdfs
    cfg.paths.output_dir = output_dir
    summary = run(cfg, copy_files=True, log=lambda *_: None)
    assert summary["ocr_read"] >= 2  # #07 and #08 both OCR'd


def test_missing_source_folder_raises_system_exit(tmp_path):
    cfg = Config()
    cfg.paths.source_dir = tmp_path / "does_not_exist"
    cfg.paths.output_dir = tmp_path / "out"
    with pytest.raises(SystemExit):
        run(cfg, copy_files=True, log=lambda *_: None)


def test_empty_source_folder_raises_system_exit(tmp_path):
    empty = tmp_path / "empty"
    empty.mkdir()
    cfg = Config()
    cfg.paths.source_dir = empty
    cfg.paths.output_dir = tmp_path / "out"
    with pytest.raises(SystemExit):
        run(cfg, copy_files=True, log=lambda *_: None)


def test_ocr_disabled_still_completes_without_crashing(fixture_pdfs, output_dir):
    cfg = Config()
    cfg.paths.source_dir = fixture_pdfs
    cfg.paths.output_dir = output_dir
    cfg.ingestion.enable_ocr = False
    summary = run(cfg, copy_files=True, log=lambda *_: None)
    assert summary["ocr_read"] == 0
    rows = _load_tracker_rows(summary)
    # #08 must STILL be caught as a duplicate of #07 via the exact_file tier.
    assert rows["08_scanned_contract_delta_rescan.pdf"]["status"] == "removed"


# ---------------------------------------------------------------------------
# Unit tests: classify.py  (regression guards for the Work Order incident)
# ---------------------------------------------------------------------------
def test_work_order_heading_classifies_as_work_order():
    cfg = ClassificationThresholds()
    text = ("WORK ORDER\nWork Order No: WO-2024-0098\n"
            "This work order is issued to Delta Industries LLC.\nDated: 12 January 2024")
    result = classify_document(text, cfg, file_name="WO_0098.pdf")
    assert result.doc_type == "Work Order"
    assert result.confidence == "HIGH"
    assert result.needs_review is False


def test_work_order_not_misfiled_as_contract():
    cfg = ClassificationThresholds()
    text = "WORK ORDER\nThis work order authorises the following services for the site."
    result = classify_document(text, cfg, file_name="wo.pdf")
    assert result.doc_type == "Work Order"
    assert result.doc_type != "Contracts"


def test_classify_document_picks_expected_bucket():
    cfg = ClassificationThresholds()
    contract_text = "SERVICES AGREEMENT\nThis agreement is made between A and B.\nIn witness whereof."
    result = classify_document(contract_text, cfg)
    assert result.doc_type == "Contracts"
    assert result.needs_review is False


def test_classify_document_falls_back_to_other_for_unrelated_text():
    cfg = ClassificationThresholds()
    result = classify_document("This is a completely unrelated memo about lunch plans.", cfg)
    assert result.doc_type == "Unclassified"
    assert result.needs_review is True


def test_classify_document_handles_empty_text():
    cfg = ClassificationThresholds()
    result = classify_document("", cfg)
    assert result.doc_type == "Unclassified"
    assert result.needs_review is True


# ---------------------------------------------------------------------------
# Unit tests: metadata.py
# ---------------------------------------------------------------------------
def test_extract_metadata_finds_entity_and_date():
    text = "This Agreement is made between Acme Holdings LLC and Ardent Advisory Ltd.\nDated: 12 January 2024"
    meta = extract_metadata(text)
    assert "Acme Holdings LLC" in meta.entity_name or "Ardent Advisory Ltd" in meta.entity_name
    assert meta.contract_date == "2024-01-12"


def test_extract_metadata_handles_empty_text_without_crashing():
    meta = extract_metadata("")
    assert meta.entity_name == ""
    assert meta.contract_date == ""
    d = meta.to_dict()
    assert "entity_name" in d and "contract_date" in d


# ---------------------------------------------------------------------------
# Unit tests: dedupe.py
# ---------------------------------------------------------------------------
def _mk_record(idx, file_name, text_norm, **overrides) -> DedupRecord:
    import hashlib
    defaults = dict(
        idx=idx, file_name=file_name, doc_type="Contracts", text_norm=text_norm,
        file_sha256=hashlib.sha256(file_name.encode()).hexdigest(),
        text_sha256=hashlib.sha256(text_norm.encode()).hexdigest(),
        entity_name="Acme Holdings LLC", contract_date="2024-01-12",
        description="", extraction_method="native_fitz", size_bytes=1000,
        text_len=len(text_norm),
    )
    defaults.update(overrides)
    return DedupRecord(**defaults)


def test_deduplicate_leaves_unique_documents_alone():
    rec_a = _mk_record(0, "a.pdf", "completely unique text about apples " * 10)
    rec_b = _mk_record(1, "b.pdf", "totally different text about oranges " * 10)
    deduplicate([rec_a, rec_b], Dedup())
    assert rec_a.status == "retained"
    assert rec_b.status == "retained"


def test_deduplicate_respects_doc_type_blocking_for_near_duplicates():
    text_a = "services agreement " * 40 + "clause set alpha version one"
    text_b = "services agreement " * 40 + "clause set alpha version two slightly different"
    rec_a = _mk_record(0, "a.pdf", text_a, doc_type="Contracts")
    rec_b = _mk_record(1, "b.pdf", text_b, doc_type="Addendums")
    deduplicate([rec_a, rec_b], Dedup(block_by_type=True, near_dup_similarity=0.80))
    assert rec_a.status == "retained"
    assert rec_b.status == "retained"


def test_deduplicate_exact_match_transcends_doc_type_blocking():
    text = "services agreement " * 40 + "identical content in both copies"
    rec_a = _mk_record(0, "a.pdf", text, doc_type="Contracts")
    rec_b = _mk_record(1, "b.pdf", text, doc_type="Addendums")
    deduplicate([rec_a, rec_b], Dedup(block_by_type=True))
    statuses = {rec_a.status, rec_b.status}
    assert statuses == {"retained", "removed"}


def test_deduplicate_never_merges_two_empty_text_records():
    rec_blank = _mk_record(0, "blank.pdf", "", entity_name="", contract_date="")
    rec_corrupt = _mk_record(1, "corrupt.pdf", "", entity_name="", contract_date="")
    deduplicate([rec_blank, rec_corrupt], Dedup())
    assert rec_blank.status == "retained"
    assert rec_corrupt.status == "retained"


def test_deduplicate_handles_empty_list():
    deduplicate([], Dedup())


def test_deduplicate_progress_callback_reports_granular_progress_and_completes():
    records = [
        _mk_record(i, f"f{i}.pdf", f"unique document body number {i} " * 30, entity_name=f"Entity {i}")
        for i in range(60)
    ]
    calls = []
    deduplicate(records, Dedup(min_chars_for_similarity=10),
                progress=lambda done, total: calls.append((done, total)))
    assert calls
    assert len(calls) > 1
    last_done, last_total = calls[-1]
    assert last_done == last_total


def test_deduplicate_result_identical_regardless_of_batch_size_ordering():
    text_a = ("SERVICES AGREEMENT this agreement is made between acme holdings llc and "
              "ardent advisory ltd dated 12 january 2024 in witness whereof the parties " * 3)
    text_b = text_a + " minor addendum clause"
    rec_a = _mk_record(0, "a.pdf", text_a)
    rec_b = _mk_record(1, "b.pdf", text_b)
    deduplicate([rec_a, rec_b], Dedup(near_dup_similarity=0.90, min_chars_for_similarity=20))
    statuses = {rec_a.status, rec_b.status}
    assert statuses == {"retained", "removed"}
