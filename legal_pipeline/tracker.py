"""tracker.py — writes the migration_tracker.xlsx report."""
from __future__ import annotations
import csv
import re
from pathlib import Path
from typing import List, Optional, Callable
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.worksheet.table import Table, TableStyleInfo
try:
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE as _ILLEGAL_XML_RE
except Exception:
    _ILLEGAL_XML_RE = re.compile("[\x00-\x08\x0b\x0c\x0e-\x1f]")

_TRACKER_COLUMNS = [
    "idx", "file_name", "doc_type", "needs_review", "status",
    "is_duplicate", "duplicate_of", "entity_name", "contract_date",
    "page_count", "is_scanned",
]
_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_REVIEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_REMOVED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_EXCLUDED_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")


def _sanitize(value):
    if isinstance(value, str):
        return _ILLEGAL_XML_RE.sub("", value)
    return value


def _row_values(r: dict, cols: List[str]) -> list:
    values = []
    for c in cols:
        v = r.get(c, "")
        if isinstance(v, (list, tuple)):
            v = ", ".join(str(x) for x in v)
        values.append(_sanitize(v))
    return values


def _write_tracker_sheet(ws, rows: List[dict], log: Callable[[str], None]):
    cols = _TRACKER_COLUMNS
    ws.append([_sanitize(c) for c in cols])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    problem_rows = 0
    for r in rows:
        values = _row_values(r, cols)
        try:
            ws.append(values)
        except IllegalCharacterError:
            problem_rows += 1
            safe_values = ["[content removed - contained unsupported characters]"
                          if isinstance(v, str) and _ILLEGAL_XML_RE.search(v) else v
                          for v in values]
            try:
                ws.append(safe_values)
            except Exception:
                ws.append([r.get("file_name", "?"), "[row could not be written]"] + [""] * (len(cols) - 2))
    if problem_rows:
        log(f"[TRACKER] Note: {problem_rows} row(s) contained characters unsupported "
            f"by Excel and had that content replaced with a placeholder.")
    for row_idx, r in enumerate(rows, start=2):
        fill = None
        if r.get("status") == "excluded":
            fill = _EXCLUDED_FILL
        elif r.get("status") == "removed":
            fill = _REMOVED_FILL
        elif r.get("needs_review"):
            fill = _REVIEW_FILL
        if fill:
            for col_idx in range(1, len(cols) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
    for col_idx, c in enumerate(cols, start=1):
        max_len = max([len(c)] + [len(str(r.get(c, ""))) for r in rows[:500]])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 50)
    ws.freeze_panes = "A2"
    if rows:
        end_col = get_column_letter(len(cols))
        table_ref = f"A1:{end_col}{len(rows) + 1}"
        table = Table(displayName="TrackerTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False
        )
        ws.add_table(table)


def _write_summary_sheet(ws, rows: List[dict], ocr_note: str,
                          consolidated_dir: Optional[Path]):
    total = len(rows)
    retained = sum(1 for r in rows if r.get("status") == "retained")
    removed = sum(1 for r in rows if r.get("status") == "removed")
    excluded = sum(1 for r in rows if r.get("status") == "excluded")
    needs_review = sum(1 for r in rows if r.get("needs_review"))
    scanned = sum(1 for r in rows if r.get("is_scanned"))
    failed = sum(1 for r in rows if r.get("extraction_method") == "failed")
    ws.append([_sanitize("Legal Document Migration — Summary")])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    pairs = [
        ("Total documents scanned", total),
        ("Retained (to migrate)", retained),
        ("Removed as duplicates", removed),
        ("Excluded — DO NOT RETAIN", excluded),
        ("Flagged for manual review", needs_review),
        ("Scanned pages (OCR candidates)", scanned),
        ("Failed extraction", failed),
        ("OCR engine status", ocr_note),
        ("Consolidated folder", str(consolidated_dir) if consolidated_dir else ""),
    ]
    for label, value in pairs:
        ws.append([_sanitize(label), _sanitize(value)])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 60


def write_rows_csv(rows: List[dict], csv_path: Path) -> Path:
    csv_path = Path(csv_path)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        cols = []
    else:
        seen = set()
        cols = []
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    cols.append(k)
                    seen.add(k)
    with csv_path.open("w", newline="", encoding="utf-8", errors="replace") as f:
        writer = csv.writer(f)
        writer.writerow(cols)
        for r in rows:
            row_values = []
            for c in cols:
                v = r.get(c, "")
                if isinstance(v, (list, tuple)):
                    v = ", ".join(str(x) for x in v)
                row_values.append(v)
            writer.writerow(row_values)
    return csv_path


def build_tracker(rows: List[dict], tracker_path: Path,
                   ocr_note: str = "", consolidated_dir: Optional[Path] = None,
                   log: Optional[Callable[[str], None]] = None) -> Path:
    log = log or (lambda *_: None)
    tracker_path = Path(tracker_path)
    tracker_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb = Workbook()
        summary_ws = wb.active
        summary_ws.title = "Summary"
        _write_summary_sheet(summary_ws, rows, ocr_note, consolidated_dir)
        tracker_ws = wb.create_sheet("Tracker")
        _write_tracker_sheet(tracker_ws, rows, log=log)
        wb.active = 0
        wb.save(str(tracker_path))
        return tracker_path
    except Exception as e:
        log(f"[TRACKER] WARNING: could not write the Excel tracker ({e}). "
            f"Falling back to a plain CSV so your results are not lost.")
        fallback_path = tracker_path.with_suffix(".csv")
        try:
            write_rows_csv(rows, fallback_path)
            log(f"[TRACKER] Fallback CSV written successfully: {fallback_path}")
            return fallback_path
        except Exception as e2:
            log(f"[TRACKER] ERROR: the fallback CSV also failed ({e2}). "
                f"Results could not be saved to a tracker file at all.")
            raise
